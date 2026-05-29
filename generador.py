import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from io import BytesIO


def generar_archivo_entrada(file_conceptos, file_empresas, file_afp, file_salud, file_mutuales, file_cajas):
    # ── Leer archivos ─────────────────────────────────────────────────────────
    conceptos_df = pd.read_excel(file_conceptos, header=1)
    empresas_df  = pd.read_excel(file_empresas, header=1)
    afp_df       = pd.read_excel(file_afp)
    salud_df     = pd.read_excel(file_salud)
    mutuales_df  = pd.read_excel(file_mutuales)
    cajas_df     = pd.read_excel(file_cajas)

    # ── Extraer listas ────────────────────────────────────────────────────────
    hab_afecto = conceptos_df[conceptos_df['Tipo']=='Haber afecto']['Nombre'].tolist()
    hab_exento = conceptos_df[conceptos_df['Tipo']=='Haber exento']['Nombre'].tolist()
    descuentos = conceptos_df[conceptos_df['Tipo']=='Descuento']['Nombre'].tolist()

    nombres_empresa = empresas_df['Nombre'].dropna().tolist()
    nombres_afp     = afp_df['nombre_afp'].dropna().tolist()
    nombres_salud   = salud_df['nombre_inst'].dropna().tolist()
    nombres_mutual  = mutuales_df['Nombre Institución'].dropna().tolist()
    nombres_caja    = cajas_df['Nombre Institución'].dropna().tolist()

    # ── Columnas fijas ────────────────────────────────────────────────────────
    FIXED = ['mes_Proceso', 'rut_trabajador', 'num_contrato', 'nombre_emp', 'id_empresa']

    DESC_LEGALES = [
        'Cotizacion AFP', 'id_afp', 'AFP Reliq meses anteriores', 'Ahorro AFP',
        'Cotizacion SALUD', 'id_salud', 'SALUD Reliq meses anteriores',
        'Seguro de Cesantia', 'CESANTIA Reliq meses anteriores',
        'Trabajo Pesado Empleado', 'APVI Ahorro voluntario mensual',
        'APVC Ahorro voluntario colectivo', 'APVI Deposito Convenido',
        'Afiliado Voluntario Cotizacion', 'Afiliado Voluntario Ahorro',
        'Prestamo solidario Remuneracion', 'Impuesto mensual',
        'IMPUESTO Reliq meses anteriores', 'Impuesto Agricola',
        'Mayor retencion Solicitada', 'Trabajo pesado Empl Reliq anteriores'
    ]

    APORTES_EMP = [
        'Trabajo Pesado', 'Trabajo pesado Reliq anteriores',
        'APVC - Aporte Empleador', 'Aporte a CCAF', 'id_ccaf',
        'CCAF Reliq meses anteriores', 'Aporte Seguro Salud',
        'Mutual', 'id_mutual', 'MUTUAL Reliq meses anteriores',
        'Seguro Invalidez y Sobrevivencia', 'SIS Reliq meses anteriores',
        'Seguro de Cesantia CI', 'CESANTIA CI Reliq meses anteriores',
        'Seguro de Cesantia Solidario', 'CESANTIA SOL Reliq meses anteriores',
        'Aporte AFP Empleador', 'Aporte AFP Empleador Reliq meses anteriores',
        'Aporte FAPP Compensación Expectativa de Vida', 'Aporte por Seguro Covid'
    ]

    all_cols = FIXED + hab_afecto + hab_exento + DESC_LEGALES + descuentos + APORTES_EMP

    # ── Meses 2020-01 a 2050-12 ───────────────────────────────────────────────
    meses = [f"{y}-{m:02d}" for y in range(2020, 2051) for m in range(1, 13)]

    # ── Workbook ──────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Liquidaciones"

    # ── Hojas ocultas ─────────────────────────────────────────────────────────
    def hidden_sheet(name, values, col2=None):
        sh = wb.create_sheet(name)
        for i, v in enumerate(values, 1):
            sh.cell(row=i, column=1, value=v)
            if col2:
                sh.cell(row=i, column=2, value=col2[i-1])
        sh.sheet_state = 'hidden'
        return sh

    hidden_sheet("lst_meses", meses)
    # empresas: col A = Nombre, col B = Empresa(id)
    sh_emp = wb.create_sheet("lst_empresas")
    for i, row in empresas_df.iterrows():
        sh_emp.cell(row=i+1, column=1, value=row['Nombre'])
        sh_emp.cell(row=i+1, column=2, value=row['Empresa'])
    sh_emp.sheet_state = 'hidden'
    n_emp = len(empresas_df)

    hidden_sheet("lst_afp", nombres_afp)
    hidden_sheet("lst_salud", nombres_salud)
    hidden_sheet("lst_mutual", nombres_mutual)
    hidden_sheet("lst_caja", nombres_caja)

    ws = wb["Liquidaciones"]

    # ── Estilos ───────────────────────────────────────────────────────────────
    FILLS = {
        'fixed':   PatternFill('solid', start_color='2E4057'),
        'hafecto': PatternFill('solid', start_color='1B6CA8'),
        'hexento': PatternFill('solid', start_color='2E8B57'),
        'dlegal':  PatternFill('solid', start_color='8B4513'),
        'desc':    PatternFill('solid', start_color='B8522A'),
        'aporte':  PatternFill('solid', start_color='6A0DAD'),
        'id':      PatternFill('solid', start_color='444444'),
    }
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
    align_c = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ID_COLS = {'id_afp', 'id_salud', 'id_ccaf', 'id_mutual', 'id_empresa'}

    def get_fill(col):
        if col in ID_COLS: return FILLS['id']
        if col in FIXED:   return FILLS['fixed']
        if col in hab_afecto: return FILLS['hafecto']
        if col in hab_exento: return FILLS['hexento']
        if col in DESC_LEGALES: return FILLS['dlegal']
        if col in descuentos:   return FILLS['desc']
        if col in APORTES_EMP:  return FILLS['aporte']
        return FILLS['fixed']

    col_map = {name: idx+1 for idx, name in enumerate(all_cols)}

    def cl(name): return get_column_letter(col_map[name])

    # ── Encabezados ───────────────────────────────────────────────────────────
    for idx, col in enumerate(all_cols, 1):
        cell = ws.cell(row=1, column=idx, value=col)
        cell.font = header_font
        cell.fill = get_fill(col)
        cell.alignment = align_c
        ws.column_dimensions[get_column_letter(idx)].width = 16

    ws.row_dimensions[1].height = 38
    ws.freeze_panes = 'A2'

    DATA_ROWS = 1000

    # ── Fórmula id_empresa ────────────────────────────────────────────────────
    emp_col = cl('nombre_emp')
    id_emp_idx = col_map['id_empresa']
    for row in range(2, DATA_ROWS + 1):
        ws.cell(row=row, column=id_emp_idx).value = (
            f'=IFERROR(VLOOKUP({emp_col}{row},lst_empresas!$A$1:$B${n_emp},2,0),"")'
        )

    # ── Validaciones ──────────────────────────────────────────────────────────
    def add_dv(ws, dv, col_name):
        dv.sqref = f"{cl(col_name)}2:{cl(col_name)}{DATA_ROWS}"
        ws.add_data_validation(dv)

    add_dv(ws, DataValidation(type="list", formula1=f"lst_meses!$A$1:$A${len(meses)}", allow_blank=True,
                               showErrorMessage=True, error="Seleccione período válido (ej: 2024-01)", errorTitle="Período inválido"), 'mes_Proceso')

    add_dv(ws, DataValidation(type="textLength", operator="between", formula1="9", formula2="10",
                               allow_blank=True, showErrorMessage=True, error="RUT debe tener 9-10 caracteres (ej: 12345678-9)", errorTitle="RUT inválido"), 'rut_trabajador')

    add_dv(ws, DataValidation(type="whole", operator="greaterThan", formula1="0",
                               allow_blank=True, showErrorMessage=True, error="Debe ser número entero mayor que 0", errorTitle="Contrato inválido"), 'num_contrato')

    add_dv(ws, DataValidation(type="list", formula1=f"lst_empresas!$A$1:$A${n_emp}", allow_blank=True,
                               showErrorMessage=True, error="Seleccione empresa de la lista", errorTitle="Empresa inválida"), 'nombre_emp')

    add_dv(ws, DataValidation(type="list", formula1=f"lst_afp!$A$1:$A${len(nombres_afp)}", allow_blank=True,
                               showErrorMessage=True, error="Seleccione AFP de la lista", errorTitle="AFP inválida"), 'id_afp')

    add_dv(ws, DataValidation(type="list", formula1=f"lst_salud!$A$1:$A${len(nombres_salud)}", allow_blank=True,
                               showErrorMessage=True, error="Seleccione institución de salud", errorTitle="Salud inválida"), 'id_salud')

    add_dv(ws, DataValidation(type="list", formula1=f"lst_mutual!$A$1:$A${len(nombres_mutual)}", allow_blank=True,
                               showErrorMessage=True, error="Seleccione mutual", errorTitle="Mutual inválida"), 'id_mutual')

    add_dv(ws, DataValidation(type="list", formula1=f"lst_caja!$A$1:$A${len(nombres_caja)}", allow_blank=True,
                               showErrorMessage=True, error="Seleccione caja de compensación", errorTitle="CCAF inválida"), 'id_ccaf')

    # Validación numérica para columnas de montos
    skip = {'mes_Proceso','rut_trabajador','num_contrato','nombre_emp','id_empresa','id_afp','id_salud','id_mutual','id_ccaf'}
    for col in all_cols:
        if col not in skip:
            dv = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True, showErrorMessage=False)
            dv.sqref = f"{cl(col)}2:{cl(col)}{DATA_ROWS}"
            ws.add_data_validation(dv)

    # ── Guardar en memoria ────────────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
