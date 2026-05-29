import pandas as pd
import numpy as np
from io import BytesIO
import calendar


# ── Mapeos de conceptos ───────────────────────────────────────────────────────

CONCEPTOS_AFP_INST = {
    'Cotizacion AFP', 'AFP Reliq meses anteriores', 'Ahorro AFP',
    'Seguro de Cesantia', 'CESANTIA Reliq meses anteriores',
    'Trabajo Pesado Empleado', 'APVI Ahorro voluntario mensual',
    'APVC Ahorro voluntario colectivo', 'APVI Deposito Convenido',
    'Afiliado Voluntario Cotizacion', 'Afiliado Voluntario Ahorro',
    'Trabajo pesado Empl Reliq anteriores', 'Seguro Invalidez y Sobrevivencia',
    'SIS Reliq meses anteriores', 'CESANTIA CI Reliq meses anteriores',
    'CESANTIA SOL Reliq meses anteriores', 'Seguro de Cesantia CI',
    'Aporte AFP Empleador', 'Aporte AFP Empleador Reliq meses anteriores',
    'Aporte FAPP Compensación Expectativa de Vida',
    'Seguro de Cesantia Solidario', 'Trabajo Pesado',
    'Trabajo pesado Reliq anteriores', 'APVC - Aporte Empleador'
}

CONCEPTOS_SALUD_INST = {'Cotizacion SALUD', 'SALUD Reliq meses anteriores', 'Aporte Seguro Salud'}
CONCEPTOS_MUTUAL_INST = {'Mutual', 'MUTUAL Reliq meses anteriores'}
CONCEPTOS_CCAF_INST = {'Aporte a CCAF', 'CCAF Reliq meses anteriores'}

CONCEPTOS_AFP_AFECTO = {
    'afp', 'reliquidaAfp', 'trabajoPesaEmpl', 'reliquidaTrabEmpl',
    'trabajoPesa', 'reliquidaTrabPesa', 'mutual', 'reliquidaMutual',
    'sis', 'reliquidaSis', 'aporteAFPemp', 'reliquidaAporteAFP',
    'aporteFAPPCEV', 'aporteFAPPBAC'
}

CONCEPTOS_CES_AFECTO = {
    'cesEmpleado', 'reliquidaCesEmpl', 'cesAporteCi',
    'reliquidaCesCi', 'cesAporteSol', 'reliquidaCesSol'
}


def es_bisiesto(anio):
    return calendar.isleap(int(anio))


def dias_mes(mes_proc):
    """Retorna el número de días del mes según mes_proc (ej: '2024-03')"""
    partes = str(mes_proc).split('-')
    anio, mes = int(partes[0]), int(partes[1])
    if mes in [1, 3, 5, 7, 8, 10, 12]:
        return 31
    elif mes in [4, 6, 9, 11]:
        return 30
    elif mes == 2:
        return 29 if es_bisiesto(anio) else 28
    return 30


def safe_float(val):
    try:
        if pd.isna(val):
            return 0.0
        return float(val)
    except:
        return 0.0


def procesar_liquidaciones(file_entrada, file_params, file_afp, file_salud, file_mutuales, file_cajas, file_empresas):
    # ── Leer archivos de referencia ───────────────────────────────────────────
    afp_df      = pd.read_excel(file_afp).set_index('nombre_afp')
    salud_df    = pd.read_excel(file_salud).set_index('nombre_inst')
    mutuales_df = pd.read_excel(file_mutuales).set_index('Nombre Institución')
    cajas_df    = pd.read_excel(file_cajas).set_index('Nombre Institución')
    empresas_df = pd.read_excel(file_empresas, header=1).set_index('Nombre')
    params_df   = pd.read_excel(file_params).set_index('mes_Proc')
    params_df.index = params_df.index.astype(str)

    entrada_df  = pd.read_excel(file_entrada)

    # ── Identificar columnas de conceptos ─────────────────────────────────────
    COLS_FIJAS = {'mes_Proceso', 'rut_trabajador', 'num_contrato', 'nombre_emp', 'id_empresa',
                  'id_afp', 'id_salud', 'id_mutual', 'id_ccaf'}
    concepto_cols = [c for c in entrada_df.columns if c not in COLS_FIJAS]

    # Leer lista de conceptos del archivo de entrada para clasificar tipos
    # Los conceptos son identificados por Nombre en el archivo de entrada
    # Necesitamos saber cuáles son Haber afecto y Haber exento
    # Usaremos los nombres tal como están en el archivo

    filas_salida = []

    for _, row in entrada_df.iterrows():
        mes_proc    = str(row.get('mes_Proceso', ''))
        rut         = row.get('rut_trabajador', '')
        num_cont    = row.get('num_contrato', '')
        nombre_emp  = row.get('nombre_emp', '')
        id_empresa  = row.get('id_empresa', '')
        id_afp_val  = row.get('id_afp', '')
        id_salud_val = row.get('id_salud', '')
        id_mutual_val = row.get('id_mutual', '')
        id_ccaf_val = row.get('id_ccaf', '')

        # Obtener parámetros del mes
        params_row = params_df.loc[mes_proc] if mes_proc in params_df.index else None

        tope_afp    = safe_float(params_row['topeImp_pesos_afp']) if params_row is not None else 0
        tope_ces    = safe_float(params_row['topeCes_pesos']) if params_row is not None else 0
        tope_salud  = safe_float(params_row['topeSalud_pesos']) if params_row is not None else 0

        # Calcular total haberes afectos y exentos
        # Necesitamos identificar qué columnas son haber afecto y exento
        # Las identificamos por su posición en el archivo (vienen antes de Cotizacion AFP)
        # Para mayor robustez, usamos todos los conceptos numéricos que no sean descuentos legales
        DESC_LEGALES_NOMBRES = {
            'Cotizacion AFP', 'AFP Reliq meses anteriores', 'Ahorro AFP',
            'Cotizacion SALUD', 'SALUD Reliq meses anteriores',
            'Seguro de Cesantia', 'CESANTIA Reliq meses anteriores',
            'Trabajo Pesado Empleado', 'APVI Ahorro voluntario mensual',
            'APVC Ahorro voluntario colectivo', 'APVI Deposito Convenido',
            'Afiliado Voluntario Cotizacion', 'Afiliado Voluntario Ahorro',
            'Prestamo solidario Remuneracion', 'Impuesto mensual',
            'IMPUESTO Reliq meses anteriores', 'Impuesto Agricola',
            'Mayor retencion Solicitada', 'Trabajo pesado Empl Reliq anteriores'
        }
        APORTES_NOMBRES = {
            'Trabajo Pesado', 'Trabajo pesado Reliq anteriores',
            'APVC - Aporte Empleador', 'Aporte a CCAF',
            'CCAF Reliq meses anteriores', 'Aporte Seguro Salud',
            'Mutual', 'MUTUAL Reliq meses anteriores',
            'Seguro Invalidez y Sobrevivencia', 'SIS Reliq meses anteriores',
            'Seguro de Cesantia CI', 'CESANTIA CI Reliq meses anteriores',
            'Seguro de Cesantia Solidario', 'CESANTIA SOL Reliq meses anteriores',
            'Aporte AFP Empleador', 'Aporte AFP Empleador Reliq meses anteriores',
            'Aporte FAPP Compensación Expectativa de Vida', 'Aporte por Seguro Covid'
        }

        # Conceptos que vienen antes de Cotizacion AFP son haberes
        cotizacion_afp_idx = concepto_cols.index('Cotizacion AFP') if 'Cotizacion AFP' in concepto_cols else len(concepto_cols)
        haberes_cols = concepto_cols[:cotizacion_afp_idx]

        total_hab_afecto = sum(safe_float(row.get(c, 0)) for c in haberes_cols
                               if c not in {'Colación', 'Movilización', 'colacion', 'movilizacion'})
        total_hab_exento = sum(safe_float(row.get(c, 0)) for c in haberes_cols
                               if c in {'Colación', 'Movilización', 'colacion', 'movilizacion'})

        # Valores clave para cálculos
        val_afp          = safe_float(row.get('Cotizacion AFP', 0))
        val_ces_empleado = safe_float(row.get('Seguro de Cesantia', 0))
        val_trab_pesa_empl = safe_float(row.get('Trabajo Pesado Empleado', 0))
        val_isapre       = safe_float(row.get('Cotizacion SALUD', 0))
        val_sueldo_base  = safe_float(row.get('Sueldo Base', 0))
        val_licencia     = safe_float(row.get('licenciaDias', 0)) if 'licenciaDias' in row.index else 0

        # Días trabajados
        dias_lic = val_licencia if val_licencia > 0 else 0
        total_dias = dias_mes(mes_proc)
        dias_trab = total_dias - dias_lic

        # Monto Init base
        monto_init = (val_sueldo_base / dias_trab * 30) if dias_trab > 0 else 0

        # Lookup helpers
        def lookup_afp_id():
            if pd.notna(id_afp_val) and str(id_afp_val) in afp_df.index:
                return afp_df.loc[str(id_afp_val), 'id_afp']
            return ''

        def lookup_salud_id():
            if pd.notna(id_salud_val) and str(id_salud_val) in salud_df.index:
                return salud_df.loc[str(id_salud_val), 'id_inst']
            return ''

        def lookup_mutual_id():
            if pd.notna(id_mutual_val) and str(id_mutual_val) in mutuales_df.index:
                return mutuales_df.loc[str(id_mutual_val), 'ID Institución']
            return ''

        def lookup_ccaf_id():
            if pd.notna(id_ccaf_val) and str(id_ccaf_val) in cajas_df.index:
                return cajas_df.loc[str(id_ccaf_val), 'ID Institución']
            return ''

        def lookup_param(col):
            if params_row is not None and col in params_df.columns:
                return safe_float(params_row[col])
            return 0

        # ── Generar filas por cada concepto ───────────────────────────────────
        for concepto_nombre in concepto_cols:
            monto = safe_float(row.get(concepto_nombre, 0))

            # Col 4 — Id del concepto
            id_concepto = concepto_nombre

            # Col 5 — Monto del concepto
            monto_concepto = monto

            # Col 6 — Afecto
            if concepto_nombre in {'totalesEmpl', 'Sueldo liquido'}:
                afecto = total_hab_afecto + total_hab_exento
            elif concepto_nombre in CONCEPTOS_AFP_AFECTO or concepto_nombre in {
                'Cotizacion AFP', 'AFP Reliq meses anteriores', 'Ahorro AFP',
                'Trabajo Pesado', 'Trabajo pesado Reliq anteriores',
                'Trabajo Pesado Empleado', 'Trabajo pesado Empl Reliq anteriores',
                'Mutual', 'MUTUAL Reliq meses anteriores',
                'Seguro Invalidez y Sobrevivencia', 'SIS Reliq meses anteriores',
                'Aporte AFP Empleador', 'Aporte AFP Empleador Reliq meses anteriores',
                'Aporte FAPP Compensación Expectativa de Vida',
                'APVC - Aporte Empleador'
            }:
                afecto = min(total_hab_afecto, tope_afp)
            elif concepto_nombre in CONCEPTOS_CES_AFECTO or concepto_nombre in {
                'Seguro de Cesantia', 'CESANTIA Reliq meses anteriores',
                'Seguro de Cesantia CI', 'CESANTIA CI Reliq meses anteriores',
                'Seguro de Cesantia Solidario', 'CESANTIA SOL Reliq meses anteriores',
            }:
                afecto = min(total_hab_afecto, tope_ces)
            elif concepto_nombre in {'Impuesto mensual', 'IMPUESTO Reliq meses anteriores',
                                      'impuesto', 'reliquidaImpuesto'}:
                afecto = total_hab_afecto - (val_afp + val_ces_empleado + val_trab_pesa_empl + min(val_isapre, tope_salud))
            elif concepto_nombre in {'reliquidaIsapre', 'SALUD Reliq meses anteriores'}:
                afecto = ''
            else:
                afecto = 0

            # Col 7 — Id de institución
            if concepto_nombre in CONCEPTOS_AFP_INST:
                id_inst = lookup_afp_id()
            elif concepto_nombre in CONCEPTOS_SALUD_INST:
                id_inst = lookup_salud_id()
            elif concepto_nombre in CONCEPTOS_MUTUAL_INST:
                id_inst = lookup_mutual_id()
            elif concepto_nombre in CONCEPTOS_CCAF_INST:
                id_inst = lookup_ccaf_id()
            else:
                id_inst = ''

            # Col 8 — Cotización de jubilación
            if concepto_nombre == 'Cotizacion AFP':
                cot_jub = safe_float(afp_df.loc[str(id_afp_val), 'cot_afp']) if pd.notna(id_afp_val) and str(id_afp_val) in afp_df.index else ''
            elif concepto_nombre in {'Cotizacion SALUD', 'isapre'}:
                cot_jub = monto_concepto
            elif concepto_nombre in {'cajaComp', 'Aporte a CCAF'}:
                cot_jub = lookup_param('aporte_Ccaf')
            elif concepto_nombre in {'Mutual', 'mutual'}:
                if pd.notna(nombre_emp) and str(nombre_emp) in empresas_df.index:
                    cot_jub = safe_float(empresas_df.loc[str(nombre_emp), 'Cotización Mutual'])
                else:
                    cot_jub = ''
            elif concepto_nombre in {'Seguro Invalidez y Sobrevivencia', 'sis'}:
                cot_jub = lookup_param('sis')
            elif concepto_nombre in {'Aporte AFP Empleador', 'aporteAFPemp'}:
                cot_jub = lookup_param('Aporte AFP')
            elif concepto_nombre in {'Aporte FAPP Compensación Expectativa de Vida', 'aporteFAPPCEV'}:
                cot_jub = lookup_param('Seg Social Exp vida')
            else:
                cot_jub = ''

            # Col 9 — Días de licencias
            if concepto_nombre in {'licenciaDias'} and monto > 0:
                dias_licencias = monto
            else:
                dias_licencias = 0

            # Col 10 — Días trabajados
            dias_trabajados = dias_trab

            # Col 13 — Total rebajas LLSS
            if concepto_nombre in {'impuesto', 'Impuesto mensual'}:
                total_rebajas = val_afp + val_ces_empleado + val_trab_pesa_empl + min(val_isapre, tope_salud)
            else:
                total_rebajas = 0

            if monto <= 0:
                continue

            filas_salida.append({
                'Fecha de proceso':          mes_proc,
                'Id empleado':               rut,
                'Número de contrato':        num_cont,
                'Id del concepto':           id_concepto,
                'Monto del concepto':        monto_concepto,
                'Afecto':                    afecto,
                'Id de institución':         id_inst,
                'Cotización de jubilación':  cot_jub,
                'Días de licencias':         dias_licencias,
                'Días trabajados':           dias_trabajados,
                'Fecha de aplicación':       'x',
                'Empresa':                   id_empresa,
                'Total de rebajas por LLSS': total_rebajas,
                'Rentas no gravadas':        0,
                'Rebaja por zona extrema':   0,
                'Jornada':                   'C',
                'Días de vacaciones':        '',
                'Monto Init':                round(monto_init, 2),
                'Fase':                      1,
            })

    # ── Crear DataFrame de salida ─────────────────────────────────────────────
    df_salida = pd.DataFrame(filas_salida)
    n_trabajadores = len(entrada_df)
    n_filas = len(df_salida)

    # ── Exportar a Excel ──────────────────────────────────────────────────────
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df_salida.to_excel(writer, index=False, sheet_name='Liquidaciones detalladas')
        ws = writer.sheets['Liquidaciones detalladas']
        # Formato encabezado
        from openpyxl.styles import Font, PatternFill, Alignment
        hfont = Font(name='Arial', bold=True, color='FFFFFF', size=9)
        hfill = PatternFill('solid', start_color='1a2744')
        halign = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for cell in ws[1]:
            cell.font = hfont
            cell.fill = hfill
            cell.alignment = halign
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = 'A2'
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18

    buf.seek(0)
    return buf.read(), n_filas, n_trabajadores
